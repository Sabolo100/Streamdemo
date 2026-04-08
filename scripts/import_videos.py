from __future__ import annotations

import csv
import json
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = ROOT / "Resoruces" / "streamfit_video_taxonomy_autotagged.xlsx"
CANONICAL_CSV_PATH = ROOT / "data" / "video_taxonomy_ssot.csv"
CANONICAL_JSON_PATH = ROOT / "data" / "video_taxonomy_ssot.json"
VALIDATION_REPORT_PATH = ROOT / "data" / "video_taxonomy_validation_report.json"
LEGACY_OUTPUT_PATH = ROOT / "data" / "videos.json"

ALLOWED_HOME_EQUIPMENT = {
    "bodyweight",
    "dumbbell",
    "kettlebell",
    "band",
    "trx",
    "bench_or_box",
}

GYM_SETUP_EQUIPMENT = {"barbell", "ring", "rope", "bosu", "weighted_vest"}
AMBIGUOUS_HOME_EQUIPMENT = {"sandbag", "medball", "other"}

WARMUP_KEYWORDS = {
    "warm",
    "mobility",
    "march",
    "prep",
    "thoracic",
    "reach",
}

COOLDOWN_KEYWORDS = {
    "stretch",
    "breathing",
    "release",
    "cooldown",
    "hip flexor",
    "groin",
    "thoracic",
}

ACTIVATION_KEYWORDS = {
    "bridge",
    "dead bug",
    "bird dog",
    "scap",
    "anti rotation",
    "anti-rotation",
    "hollow",
    "plank",
}

FINISHER_KEYWORDS = {
    "jump",
    "hop",
    "burpee",
    "thruster",
    "swing",
    "mountain climber",
    "high knees",
    "skater",
}

ADVANCED_RISK_KEYWORDS = {
    "snatch",
    "clean",
    "turkish get up",
    "turkish get-up",
    "pistol",
    "windmill",
    "handstand",
    "muscle up",
}

LARGE_SPACE_KEYWORDS = {
    "sprint",
    "run",
    "shuffle",
    "broad jump",
    "bear crawl",
    "crawl",
}

TITLE_EQUIPMENT_KEYWORDS = {
    "trx": "trx",
    "kettlebell": "kettlebell",
    "kettlbell": "kettlebell",
    "kettlbellel": "kettlebell",
    "dumbbell": "dumbbell",
    "dumbell": "dumbbell",
    "pallof": "band",
    "banded": "band",
    "band ": "band",
    " band": "band",
    "resistance band": "band",
    "mini band": "band",
    "gumikotel": "band",
    "gumiszalag": "band",
    "bench": "bench_or_box",
    "box": "bench_or_box",
    "step up": "bench_or_box",
    "pad": "bench_or_box",
    "doboz": "bench_or_box",
    "barbell": "barbell",
    "rud": "barbell",
    "landmine": "other",
    "plate": "other",
    "tarcsa": "other",
    "ring": "ring",
    "gyuru": "ring",
    "rope": "rope",
    "weighted vest": "weighted_vest",
    "sulymelleny": "weighted_vest",
    "sandbag": "sandbag",
    "homokzsak": "sandbag",
    "medball": "medball",
    "medicine ball": "medball",
    "bosu": "bosu",
}

GYM_SETUP_TITLE_KEYWORDS = {"landmine", "machine", "cable", "smith", "pulley"}
ANCHOR_SETUP_KEYWORDS = {
    "pull up",
    "chin up",
    "dip",
    "hanging leg raise",
    "hanging scapular",
    "bar hang",
    "muscle up",
    "fuggeszkedes",
    "fuggeszkedesben",
    "huzodzkodas",
    "logas",
    "rudon",
}
TRX_ANCHOR_EXCEPTIONS = {"trx pull up", "trx chin up", "trx dip"}
SEQUENCE_KEYWORDS = {
    "follow along",
    "followalong",
    "sequence",
    "routine",
    "flow",
    "circuit",
    "series",
}
WALL_KEYWORDS = {"wall", "falnal"}
BAND_ANCHOR_KEYWORDS = {
    "pallof",
    "pulldown",
    "pull down",
    "lehuzas",
    "face pull",
    "anti rotation",
    "anti-rotation",
}
COMMERCIAL_GYM_KEYWORDS = {
    "machine",
    "cable",
    "landmine",
    "smith",
    "pulley",
}
HANGING_KEYWORDS = {
    "hanging",
    "pull up",
    "chin up",
    "dip",
    "bar hang",
    "fuggeszkedes",
    "fuggeszkedesben",
    "huzodzkodas",
    "logas",
}

COOLDOWN_EXCLUSION_KEYWORDS = {
    "plank",
    "copenhagen",
    "hip lift",
    "bridge",
    "dead bug",
    "bird dog",
    "wall sit",
}

TUTORIAL_KEYWORDS = {"tutorial", "oktato video", "instruction", "how to", "guide"}
UPPER_PREP_KEYWORDS = {
    "thoracic",
    "wall slide",
    "shoulder",
    "scap",
    "pull apart",
    "face pull",
    "open book",
    "arm circle",
    "snow angel",
    "halo",
}
LOWER_PREP_KEYWORDS = {
    "hip flexor",
    "adductor",
    "groin",
    "ankle",
    "hamstring stretch",
    "quad stretch",
    "calf stretch",
    "glute stretch",
    "90 90",
}
CORE_PREP_KEYWORDS = {
    "dead bug",
    "bird dog",
    "pallof",
    "anti rotation",
    "anti-rotation",
    "plank",
    "hollow",
    "breathing",
    "crocodile",
}
BREATHING_KEYWORDS = {"breathing", "crocodile", "slow breath", "legzes"}
SCAPULAR_KEYWORDS = {"scap", "wall slide", "pull apart", "face pull", "retraction", "depression"}
THORACIC_KEYWORDS = {"thoracic", "open book", "t spine", "tspine"}
HIP_MOBILITY_KEYWORDS = {"hip flexor", "adductor", "groin", "90 90", "ankle", "hamstring stretch"}
PUSH_KEYWORDS = {
    "press",
    "push up",
    "pushup",
    "dip",
    "floor press",
    "chest press",
    "shoulder press",
    "overhead press",
    "filly press",
}
PULL_KEYWORDS = {
    "row",
    "pull apart",
    "pull down",
    "pulldown",
    "face pull",
    "reverse fly",
    "chin up",
    "pull up",
    "high row",
}
SQUAT_KEYWORDS = {"squat", "wall sit", "thruster"}
HINGE_KEYWORDS = {"deadlift", "rdl", "hinge", "swing", "good morning", "bridge", "thrust"}
LUNGE_KEYWORDS = {"lunge", "split squat", "step up", "step down", "curtsy"}
ANTI_ROTATION_KEYWORDS = {"pallof", "anti rotation", "anti-rotation"}
LATERAL_KEYWORDS = {"lateral", "side", "cossack", "curtsy", "adductor"}
ROTATION_KEYWORDS = {"rotation", "rotacio", "twist", "woodchop", "open book", "anti rotation", "anti-rotation"}
MULTIPLANAR_KEYWORDS = {"crawl", "bear crawl", "worlds greatest", "flow"}
HOLD_KEYWORDS = {
    "hold",
    "megtartas",
    "tartva",
    "plank",
    "wall sit",
    "falnal ules",
    "fuggeszkedes",
    "logas",
    "stretch",
    "breathing",
    "legzes",
}
TIME_KEYWORDS = {
    "walking",
    "march",
    "jog",
    "run",
    "high knees",
    "mountain climber",
    "skater",
    "jumping jack",
    "hegymaszo",
    "futas",
    "gyaloglas",
}
BALLISTIC_KEYWORDS = {"swing", "snatch", "clean", "jump", "hop", "bound", "throw", "thruster"}
INVERSION_KEYWORDS = {"handstand", "headstand", "cartwheel"}
OVERHEAD_KEYWORDS = {"overhead", "fej fole", "shoulder press"}
ALTERNATING_KEYWORDS = {"alternating", "valtott", "march"}
CONTRALATERAL_KEYWORDS = {"bird dog", "dead bug", "ellentetes", "opposite arm", "cross crawl"}
OFFSET_LOAD_KEYWORDS = {"single arm", "single side", "egykezes", "suitcase", "offset"}
SMR_KEYWORDS = {"smr", "foam roll", "roller", "henger", "hengerezes"}
REGRESSION_KEYWORDS = {
    "wall push up",
    "incline push up",
    "modified push up",
    "knee push up",
    "assisted",
    "supported",
    "bench supported",
    "wall supported",
}
ISOLATION_KEYWORDS = {
    "bicep curl",
    "tricep",
    "leg extension",
    "leg curl",
    "abduction",
    "adduction",
    "lateral raise",
    "front raise",
    "rear delt",
    "calf raise",
    "wrist curl",
}
SPECIALIST_KEYWORDS = {
    "copenhagen",
    "turkish get up",
    "turkish get-up",
    "pistol",
    "windmill",
    "muscle up",
    "dragon flag",
    "handstand",
    "landmine",
}
PROGRESSION_KEYWORDS = {
    "single leg",
    "single arm",
    "overhead",
    "bulgar",
    "lateral lunge",
    "cossack",
    "plyo",
    "jump",
    "explosive",
    "offset",
}
MOBILITY_RECOVERY_FAMILIES = {
    "hip_mobility",
    "tspine_mobility",
    "shoulder_mobility",
    "ankle_mobility",
    "breathing_reset",
}


def normalize_pipe_list(value: Any) -> list[str]:
    if value is None:
        return []

    text = str(value).strip()
    if not text or text.lower() == "none":
        return []

    return [part.strip() for part in text.split("|") if part.strip()]


def normalize_intensity(value: Any) -> str:
    text = str(value or "medium").strip().lower().replace("-", "_").replace(" ", "_")
    if text not in {"low", "low_medium", "medium", "medium_high", "high"}:
        return "medium"
    return text


def normalize_pattern(value: Any) -> str:
    text = str(value or "mixed_other").strip().lower()
    if text in {"core_flexion", "core_rotation", "core_anti_extension"}:
        return "core"
    return text or "mixed_other"


def title_contains(title: str, keywords: set[str]) -> bool:
    return any(keyword in title for keyword in keywords)


def simplify_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_text.lower().replace("-", " ").replace("/", " ").split())


def infer_equipment_from_title(title: str) -> list[str]:
    found: set[str] = set()
    for keyword, equipment in TITLE_EQUIPMENT_KEYWORDS.items():
        if keyword in title:
            found.add(equipment)
    return sorted(found)


def pipe_join(values: list[str]) -> str:
    return "|".join(values)


def merge_equipment(raw_equipment: list[str], inferred_equipment: list[str]) -> list[str]:
    merged = set(raw_equipment)
    merged.update(inferred_equipment)
    if not merged:
        return ["bodyweight"]
    return sorted(merged)


def pattern_is_mobility_family(movement_family_detailed: str) -> bool:
    return movement_family_detailed in MOBILITY_RECOVERY_FAMILIES


def derive_content_kind(title: str) -> str:
    return "tutorial" if title_contains(title, TUTORIAL_KEYWORDS) else "exercise"


def derive_asset_kind(title: str) -> str:
    return "followalong_sequence" if title_contains(title, SEQUENCE_KEYWORDS) else "exercise_single"


def derive_exercise_family(title: str, pattern: str) -> str:
    family_rules = [
        ("hip_flexor_stretch", {"hip flexor"}),
        ("adductor_stretch", {"adductor", "groin"}),
        ("thoracic_mobility", {"thoracic", "open book"}),
        ("breathing_reset", BREATHING_KEYWORDS),
        ("dead_bug", {"dead bug"}),
        ("bird_dog", {"bird dog"}),
        ("glute_bridge", {"glute bridge", "bridge"}),
        ("pallof_press", {"pallof"}),
        ("wall_sit", {"wall sit", "falnal ules"}),
        ("plank", {"plank"}),
        ("reverse_crunch", {"reverse crunch", "forditott haspres"}),
        ("crunch", {"crunch", "haspres", "felules"}),
        ("scapular_pull", {"scapula depression", "lapockasullyesztes", "scapular"}),
        ("overhead_press", {"overhead press", "fej fole nyomas"}),
        ("push_up", {"push up", "pushup"}),
        ("row", {"row", "evezes"}),
        ("pull_apart", {"pull apart"}),
        ("face_pull", {"face pull"}),
        ("wall_slide", {"wall slide"}),
        ("squat", SQUAT_KEYWORDS),
        ("lunge", LUNGE_KEYWORDS),
        ("hinge", HINGE_KEYWORDS),
    ]

    for family, keywords in family_rules:
        if title_contains(title, keywords):
            return family

    return pattern or "generic"


def derive_movement_family_detailed(title: str, pattern: str, exercise_family: str) -> str:
    if title_contains(title, SMR_KEYWORDS) and not title_contains(
        title,
        {"bird dog", "dead bug", "plank", "crunch", "push up", "row", "lunge", "squat", "press"},
    ):
        if title_contains(title, {"glute", "adductor", "hamstring", "hip", "farizom", "combkozelito"}):
            return "hip_mobility"
        if title_contains(title, {"thoracic", "upper back", "spine", "gerinc", "hat"}):
            return "tspine_mobility"
        if title_contains(title, {"shoulder", "vall"}):
            return "shoulder_mobility"
        if title_contains(title, {"ankle", "calf", "boka", "vadli"}):
            return "ankle_mobility"
        return "hip_mobility"

    if exercise_family in {"hip_flexor_stretch", "adductor_stretch"}:
        return "hip_mobility"
    if exercise_family == "thoracic_mobility":
        return "tspine_mobility"
    if exercise_family == "breathing_reset":
        return "breathing_reset"
    if exercise_family == "dead_bug":
        return "core_anti_extension"
    if exercise_family == "bird_dog":
        return "cross_crawl_pattern"
    if exercise_family == "pallof_press":
        return "core_anti_rotation"
    if exercise_family == "wall_sit":
        return "squat_supported"
    if exercise_family == "plank":
        if title_contains(title, {"side plank", "copenhagen", "oldal"}):
            return "core_anti_lateral_flexion"
        return "core_anti_extension"
    if exercise_family in {"reverse_crunch", "crunch"}:
        return "core_flexion"
    if exercise_family in {"scapular_pull", "pull_apart", "face_pull", "wall_slide"}:
        return "scapular_control"
    if exercise_family == "push_up":
        return "upper_push_horizontal"
    if exercise_family == "overhead_press":
        return "upper_push_vertical"
    if exercise_family == "row":
        return "upper_pull_horizontal"
    if exercise_family == "glute_bridge":
        return "glute_bridge"

    if title_contains(title, {"step up", "step down", "lepes"}):
        return "step_up"
    if title_contains(title, {"split squat", "bolgar"}):
        return "split_squat"
    if title_contains(title, LATERAL_KEYWORDS) and pattern == "lunge":
        return "lunge_lateral"
    if pattern == "lunge":
        return "lunge_forward_reverse"
    if title_contains(title, {"single leg", "egylabas"}) and pattern == "hinge":
        return "hinge_single_leg"
    if title_contains(title, {"hamstring curl", "leg curl"}):
        return "hamstring_curl"
    if title_contains(title, {"calf raise", "vadlieles"}):
        return "calf_raise"
    if title_contains(title, {"carry", "farmer", "suitcase"}):
        return "carry_hold"
    if title_contains(title, {"shoulder mobility", "halo"}):
        return "shoulder_mobility"
    if title_contains(title, {"ankle", "bokamobilitas"}):
        return "ankle_mobility"
    if title_contains(title, ROTATION_KEYWORDS) and pattern == "core":
        return "core_rotation"
    if title_contains(title, {"side plank", "copenhagen", "koppenhagai"}):
        return "core_anti_lateral_flexion"
    if title_contains(title, {"hollow", "dead bug"}) or pattern == "core":
        return "core_anti_extension"
    if title_contains(title, {"jump", "hop", "skater", "burpee"}):
        return "jump_plyometric"
    if title_contains(title, {"swing", "snatch", "clean", "thruster"}):
        return "ballistic_power"
    if pattern == "cardio_locomotion":
        return "low_impact_conditioning"
    if pattern == "push":
        return "upper_push_vertical" if title_contains(title, OVERHEAD_KEYWORDS) else "upper_push_horizontal"
    if pattern == "pull":
        if title_contains(title, {"pull up", "chin up", "pulldown", "pull down"}):
            return "upper_pull_vertical"
        return "upper_pull_horizontal"
    if pattern == "squat":
        return "squat_bilateral"
    if pattern == "hinge":
        return "hinge_bilateral"

    return "mixed_other"


def derive_balance_bucket(pattern: str, body_region: str, movement_family_detailed: str) -> str:
    if movement_family_detailed in {"hip_mobility", "tspine_mobility", "shoulder_mobility", "ankle_mobility", "breathing_reset"}:
        return "mobility_recovery"
    if movement_family_detailed in {"upper_push_horizontal", "upper_push_vertical"}:
        return "upper_push"
    if movement_family_detailed in {"upper_pull_horizontal", "upper_pull_vertical", "scapular_control"}:
        return "upper_pull"
    if movement_family_detailed in {"squat_bilateral", "squat_supported", "lunge_forward_reverse", "lunge_lateral", "split_squat", "step_up", "calf_raise"}:
        return "lower_knee"
    if movement_family_detailed in {"hinge_bilateral", "hinge_single_leg", "glute_bridge", "hamstring_curl"}:
        return "lower_hip"
    if movement_family_detailed in {
        "carry_hold",
        "core_anti_extension",
        "core_anti_rotation",
        "core_anti_lateral_flexion",
        "core_flexion",
        "core_rotation",
        "cross_crawl_pattern",
    }:
        return "trunk"
    if movement_family_detailed in {"low_impact_conditioning", "jump_plyometric", "ballistic_power"}:
        return "total_body"
    if body_region == "core" or pattern == "core":
        return "trunk"
    if body_region == "full_body" or pattern == "cardio_locomotion":
        return "total_body"
    if body_region == "upper_body":
        return "upper_push" if pattern == "push" else "upper_pull"
    if pattern == "hinge":
        return "lower_hip"
    return "lower_knee"


def derive_body_region(
    source_body_region: str,
    pattern: str,
    balance_bucket: str,
    movement_family_detailed: str,
) -> str:
    if balance_bucket in {"upper_push", "upper_pull"}:
        return "upper_body"
    if balance_bucket in {"lower_knee", "lower_hip"}:
        return "lower_body"
    if balance_bucket == "trunk":
        return "core"
    if balance_bucket == "mobility_recovery":
        if movement_family_detailed in {"hip_mobility", "ankle_mobility"}:
            return "lower_body"
        if movement_family_detailed in {"tspine_mobility", "shoulder_mobility"}:
            return "upper_body"
        if movement_family_detailed == "breathing_reset":
            return "core"
    if balance_bucket == "total_body" or pattern == "cardio_locomotion":
        return "full_body"
    return source_body_region


def derive_movement_class(
    title: str,
    pattern: str,
    movement_family_detailed: str,
    balance_bucket: str,
    exercise_family: str,
) -> str:
    if movement_family_detailed == "breathing_reset":
        return "recovery"
    if movement_family_detailed in MOBILITY_RECOVERY_FAMILIES:
        return "mobility"
    if movement_family_detailed in {"jump_plyometric", "ballistic_power"}:
        return "power"
    if movement_family_detailed == "low_impact_conditioning" or pattern == "cardio_locomotion":
        return "conditioning"
    if title_contains(title, HOLD_KEYWORDS):
        return "accessory"
    if exercise_family in {"wall_sit", "reverse_crunch", "crunch", "pallof_press", "dead_bug", "bird_dog", "scapular_pull", "pull_apart", "face_pull", "wall_slide"}:
        return "accessory"
    if title_contains(title, ISOLATION_KEYWORDS) or movement_family_detailed in {"hamstring_curl", "calf_raise"}:
        return "isolation"
    if movement_family_detailed in {
        "scapular_control",
        "carry_hold",
        "core_anti_extension",
        "core_anti_rotation",
        "core_anti_lateral_flexion",
        "core_flexion",
        "core_rotation",
        "cross_crawl_pattern",
        "glute_bridge",
        "squat_supported",
    }:
        return "accessory"
    if movement_family_detailed in {
        "squat_bilateral",
        "lunge_forward_reverse",
        "lunge_lateral",
        "split_squat",
        "step_up",
        "hinge_bilateral",
        "hinge_single_leg",
        "upper_push_horizontal",
        "upper_push_vertical",
        "upper_pull_horizontal",
        "upper_pull_vertical",
    }:
        return "compound"
    if balance_bucket in {"upper_push", "upper_pull", "lower_knee", "lower_hip", "total_body"} and pattern in {"squat", "lunge", "hinge", "push", "pull"}:
        return "compound"
    return "accessory"


def derive_variation_tier(
    title: str,
    movement_class: str,
    technical_gates: list[str],
    unilateral: bool,
    movement_family_detailed: str,
) -> str:
    if title_contains(title, REGRESSION_KEYWORDS):
        return "regression"
    if (
        title_contains(title, SPECIALIST_KEYWORDS)
        or any(gate in {"hanging", "inversion", "complex_skill"} for gate in technical_gates)
        or (movement_family_detailed == "mixed_other" and movement_class in {"compound", "accessory"})
    ):
        return "specialist"
    if (
        title_contains(title, PROGRESSION_KEYWORDS)
        or unilateral
        or any(gate in {"overhead", "ballistic"} for gate in technical_gates)
    ):
        return "progression"
    return "standard"


def derive_roles(
    base_roles: list[str],
    title: str,
    pattern: str,
    body_region: str,
    intensity: str,
    exercise_family: str,
    movement_family_detailed: str,
    movement_class: str,
    variation_tier: str,
) -> list[str]:
    roles = set(base_roles)

    if title_contains(title, COOLDOWN_EXCLUSION_KEYWORDS):
        roles.discard("cooldown")

    if title_contains(title, {"plank", "copenhagen", "koppenhagai", "wall sit", "thruster"}):
        roles.discard("warmup")

    if movement_family_detailed in {"core_anti_lateral_flexion", "core_flexion"}:
        roles.discard("warmup")

    if movement_family_detailed in {"jump_plyometric", "ballistic_power"}:
        roles.discard("warmup")

    if pattern == "mobility" or title_contains(title, WARMUP_KEYWORDS):
        roles.update({"warmup", "accessory"})

    if movement_family_detailed in {"hip_mobility", "tspine_mobility", "shoulder_mobility", "ankle_mobility"}:
        roles.update({"warmup", "accessory"})
        roles.discard("activation")
        roles.discard("main")
        roles.discard("finisher")

    if movement_family_detailed == "breathing_reset":
        roles.discard("main")
        roles.discard("finisher")

    if pattern == "mobility" and intensity == "low" and not title_contains(title, COOLDOWN_EXCLUSION_KEYWORDS):
        roles.add("cooldown")

    if title_contains(title, COOLDOWN_KEYWORDS) and not title_contains(title, COOLDOWN_EXCLUSION_KEYWORDS):
        roles.update({"cooldown", "warmup"})

    if title_contains(title, ACTIVATION_KEYWORDS):
        roles.update({"activation", "accessory"})

    if movement_family_detailed in {"cross_crawl_pattern", "core_anti_extension", "core_anti_rotation", "scapular_control", "glute_bridge"}:
        roles.update({"activation", "accessory"})

    if movement_family_detailed in MOBILITY_RECOVERY_FAMILIES and movement_family_detailed != "breathing_reset":
        roles.discard("activation")

    if movement_class == "compound":
        roles.add("main")

    if movement_class in {"power", "conditioning"}:
        roles.update({"main", "finisher"})

    if exercise_family in {"pallof_press", "dead_bug", "bird_dog", "scapular_pull"} and movement_family_detailed not in MOBILITY_RECOVERY_FAMILIES:
        roles.update({"activation", "accessory"})
        roles.discard("finisher")

    if exercise_family in {"wall_sit", "reverse_crunch", "crunch"}:
        roles.add("accessory")
        roles.discard("warmup")

    if exercise_family in {"pallof_press", "scapular_pull"}:
        roles.discard("main")

    if body_region == "core" or pattern == "core":
        roles.add("accessory")
        if intensity in {"low", "medium"} and movement_family_detailed not in MOBILITY_RECOVERY_FAMILIES:
            roles.add("activation")

    if pattern == "cardio_locomotion" or title_contains(title, FINISHER_KEYWORDS):
        roles.update({"finisher", "main"})

    if movement_family_detailed in {"core_anti_lateral_flexion", "core_flexion"}:
        roles.discard("warmup")

    if movement_class == "isolation":
        roles.discard("warmup")
        roles.discard("activation")
        roles.discard("main")
        roles.discard("finisher")
        roles.add("accessory")

    if movement_class == "recovery":
        roles.discard("main")
        roles.discard("finisher")
        roles.add("cooldown")

    if variation_tier == "specialist":
        roles.discard("warmup")

    if not roles:
        roles.add("main" if movement_class == "compound" else "accessory")

    return sorted(roles)


def derive_slot_details(
    title: str,
    session_roles: list[str],
    movement_family_detailed: str,
    balance_bucket: str,
    technical_gates: list[str],
) -> list[str]:
    slots: set[str] = set()

    if "warmup" in session_roles and (
        movement_family_detailed in MOBILITY_RECOVERY_FAMILIES
        or title_contains(title, WARMUP_KEYWORDS)
        or pattern_is_mobility_family(movement_family_detailed)
    ):
        slots.add("warmup_mobility")

    if "activation" in session_roles:
        if movement_family_detailed in {
            "glute_bridge",
            "hinge_bilateral",
            "hinge_single_leg",
            "squat_bilateral",
            "squat_supported",
            "split_squat",
            "lunge_forward_reverse",
            "lunge_lateral",
            "step_up",
        }:
            slots.add("activation_glute")
        if balance_bucket == "trunk" or movement_family_detailed in {"core_anti_extension", "core_anti_rotation", "cross_crawl_pattern"}:
            slots.add("activation_core")
        if movement_family_detailed == "breathing_reset":
            slots.add("activation_core")
        if movement_family_detailed in {"scapular_control", "upper_pull_horizontal", "upper_pull_vertical", "shoulder_mobility"}:
            slots.add("activation_scapula")

    if "main" in session_roles:
        if movement_family_detailed == "jump_plyometric":
            slots.add("power")
        elif movement_family_detailed in {"low_impact_conditioning", "ballistic_power"}:
            slots.add("conditioning" if movement_family_detailed == "low_impact_conditioning" else "power")
        else:
            slots.add("main_strength")

    if "accessory" in session_roles and movement_family_detailed not in MOBILITY_RECOVERY_FAMILIES:
        if any(gate in {"complex_skill", "hanging"} for gate in technical_gates):
            slots.add("skill")
        else:
            slots.add("accessory_strength")

    if "finisher" in session_roles:
        if movement_family_detailed in {"jump_plyometric", "ballistic_power"}:
            slots.add("power")
        else:
            slots.add("conditioning")

    if "cooldown" in session_roles:
        if movement_family_detailed == "breathing_reset" or title_contains(title, BREATHING_KEYWORDS):
            slots.update({"cooldown_breathing", "cooldown_downregulation"})
        elif balance_bucket == "mobility_recovery":
            slots.add("cooldown_mobility")
        else:
            slots.add("cooldown_downregulation")

    return sorted(slots)


def derive_builder_tags(
    title: str,
    pattern: str,
    body_region: str,
    session_roles: list[str],
    slot_details: list[str],
    balance_bucket: str,
    movement_family_detailed: str,
    content_kind: str,
) -> list[str]:
    if content_kind == "tutorial":
        return []

    tags: set[str] = set()

    if "warmup_mobility" in slot_details:
        if movement_family_detailed in {"hip_mobility", "ankle_mobility"} or body_region == "lower_body":
            tags.add("prep_lower")
        elif movement_family_detailed in {"tspine_mobility", "shoulder_mobility", "scapular_control"} or body_region == "upper_body":
            tags.add("prep_upper")
        elif balance_bucket == "trunk" or body_region == "core":
            tags.add("prep_core")
        else:
            tags.add("prep_full")

    if "activation_glute" in slot_details:
        tags.add("activation_lower")
    if "activation_core" in slot_details:
        tags.add("activation_core")
    if "activation_scapula" in slot_details:
        tags.add("activation_upper")
        tags.add("scapular_control")

    if balance_bucket == "upper_push":
        tags.update({"strength_upper_push", "accessory_upper", "push_pattern"})
    elif balance_bucket == "upper_pull":
        tags.update({"strength_upper_pull", "accessory_upper", "pull_pattern"})
    elif balance_bucket == "lower_knee":
        tags.add("accessory_lower")
        if movement_family_detailed in {"split_squat", "lunge_forward_reverse", "lunge_lateral", "step_up"}:
            tags.update({"strength_lower_lunge", "lunge_pattern"})
        else:
            tags.update({"strength_lower_squat", "squat_pattern"})
    elif balance_bucket == "lower_hip":
        tags.update({"strength_lower_hinge", "accessory_lower", "hinge_pattern"})
    elif balance_bucket == "trunk":
        tags.update({"strength_core", "accessory_core"})
    elif balance_bucket == "total_body" and "activation" in session_roles:
        tags.add("activation_full")
    elif balance_bucket == "total_body" and "warmup" in session_roles:
        tags.add("prep_full")

    if "cooldown_breathing" in slot_details:
        tags.update({"recovery_breathing", "recovery_core"})
    if "cooldown_mobility" in slot_details:
        if movement_family_detailed in {"hip_mobility", "ankle_mobility"}:
            tags.add("recovery_lower")
            tags.add("mobility_hips")
        elif movement_family_detailed in {"tspine_mobility", "shoulder_mobility", "scapular_control"}:
            tags.add("recovery_upper")
            if movement_family_detailed == "tspine_mobility":
                tags.add("mobility_thoracic")
        elif balance_bucket == "trunk":
            tags.add("recovery_core")
        else:
            tags.update({"recovery_upper", "recovery_lower"})

    if movement_family_detailed == "core_anti_rotation":
        tags.add("anti_rotation")
        tags.update({"activation_core", "strength_core", "accessory_core"})

    if movement_family_detailed == "scapular_control":
        tags.add("scapular_control")

    if title_contains(title, THORACIC_KEYWORDS):
        tags.add("mobility_thoracic")
    if title_contains(title, HIP_MOBILITY_KEYWORDS):
        tags.add("mobility_hips")

    return sorted(tags)


def derive_position_detail(title: str, position: str) -> str:
    if title_contains(title, {"half kneeling", "felterden"}):
        return "half_kneeling"
    if title_contains(title, {"tall kneeling", "magas terdeles"}):
        return "tall_kneeling"
    if title_contains(title, {"quadruped", "terdelotamasz"}):
        return "quadruped"
    if title_contains(title, {"plank", "alkartamasz"}):
        return "plank"
    if title_contains(title, HANGING_KEYWORDS):
        return "hanging"
    if title_contains(title, {"supine", "hanyatt", "haton fekve"}):
        return "supine"
    if title_contains(title, {"prone", "hason fekve"}):
        return "prone"
    if title_contains(title, {"side lying", "oldalfekves"}):
        return "side_lying"
    if title_contains(title, {"seated", "ulve", "ulesben"}):
        return "seated"
    if title_contains(title, WALL_KEYWORDS):
        return "wall_supported"
    if title_contains(title, {"bench supported", "padon", "fekvopadon", "boxon"}):
        return "bench_supported"
    if title_contains(title, {"staggered", "aszimmetrikus allas"}):
        return "staggered_stance"
    if title_contains(title, {"split stance", "split squat", "kitores", "bolgar"}):
        return "split_stance"
    if position == "floor":
        return "supine"
    return "standing"


def derive_contraction_style(title: str, pattern: str, movement_family_detailed: str) -> str:
    if movement_family_detailed in {"hip_mobility", "tspine_mobility", "shoulder_mobility", "ankle_mobility", "breathing_reset"}:
        return "controlled_mobility"
    if title_contains(title, HOLD_KEYWORDS) or movement_family_detailed in {"squat_supported", "carry_hold"}:
        return "isometric_hold"
    if movement_family_detailed in {"jump_plyometric", "ballistic_power"} or title_contains(title, BALLISTIC_KEYWORDS):
        return "ballistic"
    if pattern == "cardio_locomotion" or title_contains(title, TIME_KEYWORDS):
        return "cyclical_endurance"
    return "dynamic_reps"


def derive_plane_of_motion(title: str, movement_family_detailed: str, pattern: str) -> str:
    if title_contains(title, MULTIPLANAR_KEYWORDS):
        return "multiplanar"
    if title_contains(title, ROTATION_KEYWORDS) or movement_family_detailed in {"core_rotation", "core_anti_rotation", "tspine_mobility"}:
        return "transverse"
    if title_contains(title, LATERAL_KEYWORDS) or movement_family_detailed in {"lunge_lateral", "core_anti_lateral_flexion"}:
        return "frontal"
    if movement_family_detailed in {"jump_plyometric", "ballistic_power"} and pattern == "cardio_locomotion":
        return "multiplanar"
    return "sagittal"


def requires_anchor_setup(title: str, equipment: list[str]) -> bool:
    if "trx" in equipment and title_contains(title, TRX_ANCHOR_EXCEPTIONS):
        return False
    return title_contains(title, ANCHOR_SETUP_KEYWORDS)


def derive_environment_access(title: str, equipment: list[str], requires_gym_setup: bool) -> str:
    if title_contains(title, HANGING_KEYWORDS):
        return "hanging_rig_needed"
    if "trx" in equipment or any(item in equipment for item in {"ring"}) or requires_anchor_setup(title, equipment):
        return "anchor_or_rig_needed"
    if "band" in equipment and title_contains(title, BAND_ANCHOR_KEYWORDS):
        return "band_anchor_needed"
    if "bench_or_box" in equipment or title_contains(title, {"step up", "step down", "box", "bench", "pad", "doboz"}):
        return "bench_or_box_needed"
    if title_contains(title, WALL_KEYWORDS):
        return "wall_needed"
    if requires_gym_setup or title_contains(title, COMMERCIAL_GYM_KEYWORDS):
        return "commercial_gym_preferred"
    return "home_open"


def derive_requires_large_space(title: str, pattern: str) -> bool:
    return pattern == "cardio_locomotion" and title_contains(title, LARGE_SPACE_KEYWORDS)


def derive_advanced_risk(title: str, complexity: str) -> bool:
    return complexity == "complex" or title_contains(title, ADVANCED_RISK_KEYWORDS)


def derive_requires_partner(title: str) -> bool:
    return "partner" in title


def derive_requires_gym_setup(equipment: list[str], title: str) -> bool:
    return (
        any(item in GYM_SETUP_EQUIPMENT for item in equipment)
        or "machine" in title
        or title_contains(title, GYM_SETUP_TITLE_KEYWORDS)
        or requires_anchor_setup(title, equipment)
    )


def derive_technical_gates(
    title: str,
    movement_family_detailed: str,
    complexity: str,
    advanced_risk: bool,
    environment_access: str,
) -> list[str]:
    gates: set[str] = set()

    if title_contains(title, OVERHEAD_KEYWORDS) or movement_family_detailed == "upper_push_vertical":
        gates.add("overhead")
    if environment_access == "hanging_rig_needed":
        gates.add("hanging")
    if title_contains(title, INVERSION_KEYWORDS):
        gates.add("inversion")
    if movement_family_detailed in {"jump_plyometric", "ballistic_power"} or title_contains(title, BALLISTIC_KEYWORDS):
        gates.add("ballistic")
    if advanced_risk or complexity == "complex":
        gates.add("complex_skill")

    return sorted(gates) if gates else ["standard"]


def derive_limb_pattern(title: str, unilateral_flag: bool, equipment: list[str], movement_family_detailed: str) -> str:
    if movement_family_detailed == "cross_crawl_pattern" or title_contains(title, CONTRALATERAL_KEYWORDS):
        return "contralateral"
    if title_contains(title, ALTERNATING_KEYWORDS):
        return "alternating"
    if unilateral_flag and title_contains(title, OFFSET_LOAD_KEYWORDS) and any(item != "bodyweight" for item in equipment):
        return "offset_loaded"
    if unilateral_flag:
        return "unilateral"
    return "bilateral"


def derive_prescription_profile(
    movement_class: str,
    balance_bucket: str,
    movement_family_detailed: str,
) -> str:
    if movement_class == "recovery":
        return "recovery_reset"
    if movement_class == "mobility":
        return "mobility_prep"
    if movement_class == "conditioning":
        return "conditioning_interval"
    if movement_class == "power":
        return "power_output"
    if movement_class == "isolation":
        return "isolation_volume"
    if balance_bucket == "trunk" or movement_family_detailed in {
        "core_anti_extension",
        "core_anti_rotation",
        "core_anti_lateral_flexion",
        "core_flexion",
        "core_rotation",
        "cross_crawl_pattern",
    }:
        return "core_control"
    if movement_class == "compound":
        return "compound_strength"
    return "accessory_volume"


def derive_preferred_format(
    title: str,
    pattern: str,
    movement_family_detailed: str,
    contraction_style: str,
) -> str:
    if contraction_style == "isometric_hold":
        return "hold"
    if contraction_style in {"cyclical_endurance", "controlled_mobility"}:
        return "time"
    if title_contains(title, HOLD_KEYWORDS):
        return "hold"
    if title_contains(title, TIME_KEYWORDS):
        return "time"
    if pattern == "cardio_locomotion" or movement_family_detailed == "low_impact_conditioning":
        return "time"
    return "reps"


def derive_home_safe(
    equipment: list[str],
    requires_large_space: bool,
    requires_partner: bool,
    environment_access: str,
) -> bool:
    if requires_large_space or requires_partner:
        return False
    if any(item in AMBIGUOUS_HOME_EQUIPMENT for item in equipment):
        return False
    if environment_access in {"hanging_rig_needed", "commercial_gym_preferred"}:
        return False
    return all(item in ALLOWED_HOME_EQUIPMENT or item == "ring" for item in equipment)


def derive_builder_status(
    content_kind: str,
    asset_kind: str,
    beginner_friendly: str,
    advanced_risk: bool,
    technical_gates: list[str],
    environment_access: str,
    variation_tier: str,
) -> str:
    if content_kind == "tutorial" or asset_kind == "followalong_sequence":
        return "exclude"
    if advanced_risk or beginner_friendly == "no":
        return "manual_review"
    if variation_tier == "specialist":
        return "manual_review"
    if any(gate in {"hanging", "inversion", "ballistic", "complex_skill"} for gate in technical_gates):
        return "manual_review"
    if environment_access in {"hanging_rig_needed", "commercial_gym_preferred"}:
        return "manual_review"
    return "include"


def build_qa_flags(
    raw_equipment: list[str],
    inferred_equipment: list[str],
    equipment: list[str],
    title: str,
    requires_gym_setup: bool,
    home_safe: bool,
    session_roles: list[str],
    content_kind: str,
    asset_kind: str,
    builder_status: str,
) -> list[str]:
    flags: list[str] = []

    if inferred_equipment:
        flags.append(f"equipment_inferred:{pipe_join(inferred_equipment)}")
    if raw_equipment != equipment:
        flags.append("canonical_equipment_differs_from_source")
    if requires_gym_setup:
        flags.append("gym_setup_required")
    if requires_anchor_setup(title, equipment):
        flags.append("fixed_anchor_or_bar_required")
    if content_kind == "tutorial":
        flags.append("instructional_content")
    if asset_kind == "followalong_sequence":
        flags.append("followalong_asset")
    if builder_status != "include":
        flags.append(f"builder_status:{builder_status}")
    if not home_safe:
        flags.append("excluded_from_home_safe_pool")
    if "cooldown" in session_roles and title_contains(title, COOLDOWN_EXCLUSION_KEYWORDS):
        flags.append("cooldown_role_removed_by_validation")

    return flags


def validate_runtime_record(record: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    title = simplify_text(record["titleOriginal"])
    equipment = set(record["equipmentTypes"])
    builder_tags = set(record["builderTags"])
    slot_details = set(record["slotDetails"])
    technical_gates = set(record["technicalGates"])

    if "landmine" in title and record["homeSafe"]:
        errors.append("landmine_item_marked_home_safe")
    if ("banded" in title or "gumikotel" in title) and "band" not in equipment:
        errors.append("band_keyword_missing_band_equipment")
    if requires_anchor_setup(title, list(equipment)) and record["environmentAccess"] not in {"anchor_or_rig_needed", "hanging_rig_needed"}:
        errors.append("anchor_based_exercise_missing_environment_access")
    if any(keyword in title for keyword in ["barbell", "rud", "plate", "tarcsa"]) and record["homeSafe"]:
        errors.append("loaded_barbell_or_plate_item_marked_home_safe")
    if "cooldown" in record["sessionRoleFit"] and title_contains(title, COOLDOWN_EXCLUSION_KEYWORDS):
        errors.append("cooldown_role_contains_non_cooldown_title")
    if record["homeSafe"] and any(item not in ALLOWED_HOME_EQUIPMENT and item != "ring" for item in equipment):
        errors.append("home_safe_item_contains_disallowed_equipment")
    if title_contains(title, TUTORIAL_KEYWORDS) and record["contentKind"] != "tutorial":
        errors.append("tutorial_title_missing_tutorial_content_kind")
    if record["contentKind"] == "exercise" and not record["exerciseFamily"]:
        errors.append("exercise_family_missing")
    if record["assetKind"] == "followalong_sequence" and record["builderStatus"] == "include":
        errors.append("followalong_sequence_marked_include")
    if record["contentKind"] == "tutorial" and record["builderStatus"] != "exclude":
        errors.append("tutorial_not_excluded_from_builder")
    if record["contractionStyle"] == "isometric_hold" and record["preferredFormat"] != "hold":
        errors.append("isometric_hold_missing_hold_format")
    if record["contractionStyle"] == "controlled_mobility" and record["preferredFormat"] not in {"time", "hold"}:
        errors.append("mobility_item_missing_time_or_hold_format")
    if record["movementClass"] == "compound" and record["preferredFormat"] == "hold":
        errors.append("compound_item_marked_hold")
    if record["movementClass"] == "isolation" and "main" in record["sessionRoleFit"]:
        errors.append("isolation_item_marked_main")
    if record["movementClass"] == "recovery" and "main" in record["sessionRoleFit"]:
        errors.append("recovery_item_marked_main")
    if "warmup" in record["sessionRoleFit"] and "warmup_mobility" not in slot_details:
        errors.append("warmup_role_missing_slot_detail")
    if "activation" in record["sessionRoleFit"] and not slot_details.intersection({"activation_glute", "activation_core", "activation_scapula"}):
        errors.append("activation_role_missing_slot_detail")
    if "cooldown" in record["sessionRoleFit"] and not slot_details.intersection({"cooldown_breathing", "cooldown_mobility", "cooldown_downregulation"}):
        errors.append("cooldown_role_missing_slot_detail")
    if "main" in record["sessionRoleFit"] and not slot_details.intersection({"main_strength", "conditioning", "power"}):
        errors.append("main_role_missing_slot_detail")
    if "warmup" in record["sessionRoleFit"] and not builder_tags.intersection({"prep_upper", "prep_lower", "prep_core", "prep_full"}):
        errors.append("warmup_role_missing_prep_builder_tags")
    if record["environmentAccess"] == "hanging_rig_needed" and "hanging" not in technical_gates:
        errors.append("hanging_environment_missing_hanging_gate")
    if record["balanceBucket"] == "mobility_recovery" and record["preferredFormat"] == "reps":
        errors.append("mobility_recovery_item_marked_reps")

    return errors


def write_csv(rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise ValueError("No taxonomy rows to write.")

    fieldnames = list(rows[0].keys())
    with CANONICAL_CSV_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_tags(
    equipment: list[str],
    pattern: str,
    body_region: str,
    impact: str,
    complexity: str,
    position: str,
    builder_tags: list[str],
    balance_bucket: str,
    movement_family_detailed: str,
    slot_details: list[str],
    technical_gates: list[str],
    environment_access: str,
    movement_class: str,
    variation_tier: str,
    prescription_profile: str,
) -> list[str]:
    tags = set(equipment)
    tags.update(
        {
            pattern,
            body_region,
            impact,
            complexity,
            position,
            balance_bucket,
            movement_family_detailed,
            environment_access,
            movement_class,
            variation_tier,
            prescription_profile,
        }
    )
    tags.update(builder_tags)
    tags.update(slot_details)
    tags.update(technical_gates)
    return sorted(tag for tag in tags if tag)


def main() -> None:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    sheet = workbook["AutoTaggedVideos"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}

    runtime_videos: list[dict[str, Any]] = []
    canonical_rows: list[dict[str, Any]] = []
    validation_errors: dict[int, list[str]] = {}

    for row in rows:
        raw_equipment = normalize_pipe_list(row[index["equipment_type"]])
        title = str(row[index["english_title"]] or "").strip()
        title_original = str(row[index["title_original"]] or title).strip()
        title_search = simplify_text(title_original)
        inferred_equipment = infer_equipment_from_title(title_search)
        equipment = merge_equipment(raw_equipment, inferred_equipment)
        pattern = normalize_pattern(row[index["primary_pattern"]])
        secondary_pattern = normalize_pattern(row[index["secondary_pattern"]]) if row[index["secondary_pattern"]] else None
        source_body_region = str(row[index["body_region"]] or "full_body").strip().lower()
        impact = str(row[index["impact_level"]] or "low").strip().lower()
        complexity = str(row[index["complexity_level"]] or "basic").strip().lower()
        intensity = normalize_intensity(row[index["intensity_estimate"]])
        position = str(row[index["position_type"]] or "standing").strip().lower()
        contraindications = normalize_pipe_list(row[index["contraindications"]])
        base_roles = normalize_pipe_list(row[index["session_role_fit"]])
        beginner_friendly = str(row[index["beginner_friendly"]] or "conditional").strip().lower()
        unilateral = str(row[index["unilateral_flag"]] or "no").strip().lower() == "yes"

        content_kind = derive_content_kind(title_search)
        asset_kind = derive_asset_kind(title_search)
        exercise_family = derive_exercise_family(title_search, pattern)
        movement_family_detailed = derive_movement_family_detailed(title_search, pattern, exercise_family)
        balance_bucket = derive_balance_bucket(pattern, source_body_region, movement_family_detailed)
        body_region = derive_body_region(source_body_region, pattern, balance_bucket, movement_family_detailed)
        base_requires_gym_setup = derive_requires_gym_setup(equipment, title_search)
        environment_access = derive_environment_access(title_search, equipment, base_requires_gym_setup)
        advanced_risk = derive_advanced_risk(title_search, complexity)
        technical_gates = derive_technical_gates(
            title_search,
            movement_family_detailed,
            complexity,
            advanced_risk,
            environment_access,
        )
        requires_large_space = derive_requires_large_space(title_search, pattern)
        requires_partner = derive_requires_partner(title_search)
        requires_gym_setup = base_requires_gym_setup or environment_access in {
            "anchor_or_rig_needed",
            "hanging_rig_needed",
            "commercial_gym_preferred",
        }
        position_detail = derive_position_detail(title_search, position)
        contraction_style = derive_contraction_style(title_search, pattern, movement_family_detailed)
        plane_of_motion = derive_plane_of_motion(title_search, movement_family_detailed, pattern)
        limb_pattern = derive_limb_pattern(title_search, unilateral, equipment, movement_family_detailed)
        movement_class = derive_movement_class(
            title_search,
            pattern,
            movement_family_detailed,
            balance_bucket,
            exercise_family,
        )
        variation_tier = derive_variation_tier(
            title_search,
            movement_class,
            technical_gates,
            unilateral,
            movement_family_detailed,
        )
        session_roles = derive_roles(
            base_roles,
            title_search,
            pattern,
            body_region,
            intensity,
            exercise_family,
            movement_family_detailed,
            movement_class,
            variation_tier,
        )
        slot_details = derive_slot_details(
            title_search,
            session_roles,
            movement_family_detailed,
            balance_bucket,
            technical_gates,
        )
        builder_tags = derive_builder_tags(
            title_search,
            pattern,
            body_region,
            session_roles,
            slot_details,
            balance_bucket,
            movement_family_detailed,
            content_kind,
        )
        preferred_format = derive_preferred_format(
            title_search,
            pattern,
            movement_family_detailed,
            contraction_style,
        )
        prescription_profile = derive_prescription_profile(
            movement_class,
            balance_bucket,
            movement_family_detailed,
        )
        home_safe = derive_home_safe(
            equipment,
            requires_large_space,
            requires_partner,
            environment_access,
        )
        builder_status = derive_builder_status(
            content_kind,
            asset_kind,
            beginner_friendly,
            advanced_risk,
            technical_gates,
            environment_access,
            variation_tier,
        )

        runtime_record = {
            "id": int(row[index["id"]]),
            "titleOriginal": title_original,
            "title": title,
            "titleHu": str(row[index["hungarian_title"]] or "").strip(),
            "videoUrl": str(row[index["video_url"]] or "").strip(),
            "contentKind": content_kind,
            "assetKind": asset_kind,
            "builderStatus": builder_status,
            "exerciseFamily": exercise_family,
            "movementFamilyDetailed": movement_family_detailed,
            "builderTags": builder_tags,
            "balanceBucket": balance_bucket,
            "slotDetails": slot_details,
            "preferredFormat": preferred_format,
            "equipmentTypes": equipment or ["bodyweight"],
            "primaryPattern": pattern,
            "secondaryPattern": secondary_pattern,
            "bodyRegion": body_region,
            "impactLevel": impact,
            "complexityLevel": complexity,
            "intensityEstimate": intensity,
            "unilateral": unilateral,
            "positionType": position,
            "positionDetail": position_detail,
            "contractionStyle": contraction_style,
            "planeOfMotion": plane_of_motion,
            "environmentAccess": environment_access,
            "technicalGates": technical_gates,
            "limbPattern": limb_pattern,
            "movementClass": movement_class,
            "variationTier": variation_tier,
            "prescriptionProfile": prescription_profile,
            "sessionRoleFit": session_roles,
            "beginnerFriendly": beginner_friendly,
            "contraindications": contraindications,
            "homeSafe": home_safe,
            "requiresLargeSpace": requires_large_space,
            "requiresPartner": requires_partner,
            "requiresGymSetup": requires_gym_setup,
            "advancedRisk": advanced_risk,
            "tags": build_tags(
                equipment,
                pattern,
                body_region,
                impact,
                complexity,
                position,
                builder_tags,
                balance_bucket,
                movement_family_detailed,
                slot_details,
                technical_gates,
                environment_access,
                movement_class,
                variation_tier,
                prescription_profile,
            ),
        }

        errors = validate_runtime_record(runtime_record)
        if errors:
            validation_errors[runtime_record["id"]] = errors

        canonical_rows.append(
            {
                "id": runtime_record["id"],
                "title_original": title_original,
                "english_title": title,
                "hungarian_title": runtime_record["titleHu"],
                "video_url": runtime_record["videoUrl"],
                "content_kind": content_kind,
                "asset_kind": asset_kind,
                "builder_status": builder_status,
                "exercise_family": exercise_family,
                "movement_family_detailed": movement_family_detailed,
                "builder_tags": pipe_join(builder_tags),
                "balance_bucket": balance_bucket,
                "slot_details": pipe_join(slot_details),
                "preferred_format": preferred_format,
                "source_equipment_types": pipe_join(raw_equipment),
                "inferred_equipment_types": pipe_join(inferred_equipment),
                "canonical_equipment_types": pipe_join(runtime_record["equipmentTypes"]),
                "primary_pattern": pattern,
                "secondary_pattern": secondary_pattern or "",
                "source_body_region": source_body_region,
                "body_region": body_region,
                "impact_level": impact,
                "complexity_level": complexity,
                "intensity_estimate": intensity,
                "position_type": position,
                "position_detail": position_detail,
                "contraction_style": contraction_style,
                "plane_of_motion": plane_of_motion,
                "environment_access": environment_access,
                "technical_gates": pipe_join(technical_gates),
                "limb_pattern": limb_pattern,
                "movement_class": movement_class,
                "variation_tier": variation_tier,
                "prescription_profile": prescription_profile,
                "unilateral_flag": "yes" if runtime_record["unilateral"] else "no",
                "beginner_friendly": runtime_record["beginnerFriendly"],
                "session_role_fit": pipe_join(session_roles),
                "contraindications": pipe_join(contraindications),
                "home_safe": "yes" if home_safe else "no",
                "requires_large_space": "yes" if requires_large_space else "no",
                "requires_partner": "yes" if requires_partner else "no",
                "requires_gym_setup": "yes" if requires_gym_setup else "no",
                "advanced_risk": "yes" if advanced_risk else "no",
                "qa_flags": pipe_join(
                    build_qa_flags(
                        raw_equipment,
                        inferred_equipment,
                        runtime_record["equipmentTypes"],
                        title_search,
                        requires_gym_setup,
                        home_safe,
                        session_roles,
                        content_kind,
                        asset_kind,
                        builder_status,
                    )
                ),
                "validation_errors": pipe_join(errors),
                "taxonomy_status": "validated" if not errors else "needs_review",
            }
        )

        runtime_videos.append(runtime_record)

    write_csv(canonical_rows)
    CANONICAL_JSON_PATH.write_text(
        json.dumps(runtime_videos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    LEGACY_OUTPUT_PATH.write_text(
        json.dumps(runtime_videos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    VALIDATION_REPORT_PATH.write_text(
        json.dumps(
            {
                "source_workbook": str(SOURCE_PATH),
                "canonical_csv": str(CANONICAL_CSV_PATH),
                "canonical_json": str(CANONICAL_JSON_PATH),
                "row_count": len(runtime_videos),
                "validation_error_count": len(validation_errors),
                "validation_errors": validation_errors,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Wrote canonical taxonomy CSV to {CANONICAL_CSV_PATH}")
    print(f"Wrote canonical runtime JSON to {CANONICAL_JSON_PATH}")
    print(f"Wrote validation report to {VALIDATION_REPORT_PATH}")


if __name__ == "__main__":
    main()
